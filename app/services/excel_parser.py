"""
Parser de archivos Excel (.xlsx, .xls) para Phoenix Legal.

FASE 1C: MULTI-FORMATO
Objetivo: Extraer contenido de archivos Excel de forma estructurada.

Casos de uso:
- Balances de situación
- Cuentas de Pérdidas y Ganancias
- Listados de acreedores
- Extractos bancarios

PRINCIPIOS:
- Extraer todo el texto visible
- Preservar estructura de filas/columnas
- Incluir nombre de hojas
- No interpretar ni analizar
"""
from __future__ import annotations

from typing import Dict, List, Optional
from pathlib import Path
import io

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelParseResult:
    """
    Resultado del parsing de un archivo Excel.
    
    Atributos:
        texto: Texto extraído (representación textual del contenido)
        num_paginas: Número de hojas en el libro
        tipo_documento: Siempre "excel"
        page_offsets: Diccionario con offsets de cada hoja
        sheets_content: Contenido por hoja (opcional, para análisis posterior)
    """
    
    def __init__(
        self,
        texto: str,
        num_paginas: int,
        page_offsets: Dict[int, tuple[int, int]],
        sheets_content: Optional[Dict[str, List[List[str]]]] = None,
    ):
        self.texto = texto
        self.num_paginas = num_paginas
        self.tipo_documento = "excel"
        self.page_offsets = page_offsets
        self.sheets_content = sheets_content or {}


def parse_excel_file(file_path: str) -> ExcelParseResult:
    """
    Parsea un archivo Excel y extrae su contenido como texto.
    
    Args:
        file_path: Ruta al archivo Excel (.xlsx, .xls)
        
    Returns:
        ExcelParseResult con el contenido extraído
        
    Raises:
        Exception: Si el archivo no se puede leer o no es un Excel válido
        
    Estrategia de extracción:
        1. Abrir workbook con openpyxl
        2. Iterar sobre cada hoja
        3. Extraer contenido celda por celda
        4. Generar representación textual estructurada
        5. Calcular offsets para cada hoja
    """
    # Abrir workbook (solo datos, sin fórmulas evaluadas)
    workbook = load_workbook(filename=file_path, data_only=True, read_only=True)
    
    sheets_content: Dict[str, List[List[str]]] = {}
    full_text_parts: List[str] = []
    page_offsets: Dict[int, tuple[int, int]] = {}
    
    current_offset = 0
    
    for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
        sheet: Worksheet = workbook[sheet_name]
        
        # Inicio de offset para esta hoja
        start_offset = current_offset
        
        # Encabezado de hoja
        sheet_header = f"\n{'=' * 80}\n"
        sheet_header += f"HOJA: {sheet_name}\n"
        sheet_header += f"{'=' * 80}\n\n"
        full_text_parts.append(sheet_header)
        current_offset += len(sheet_header)
        
        # Extraer contenido de la hoja
        sheet_rows: List[List[str]] = []
        
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # Convertir valores de celda a string
            row_values = [
                str(cell) if cell is not None else ""
                for cell in row
            ]
            
            # Filtrar filas completamente vacías
            if any(val.strip() for val in row_values):
                sheet_rows.append(row_values)
                
                # Añadir fila al texto (formato tabular)
                row_text = " | ".join(row_values)
                row_line = f"Fila {row_idx}: {row_text}\n"
                full_text_parts.append(row_line)
                current_offset += len(row_line)
        
        # Guardar contenido estructurado de la hoja
        sheets_content[sheet_name] = sheet_rows
        
        # Fin de offset para esta hoja
        end_offset = current_offset
        page_offsets[sheet_idx] = (start_offset, end_offset)
        
        # Separador entre hojas
        separator = "\n\n"
        full_text_parts.append(separator)
        current_offset += len(separator)
    
    workbook.close()
    
    # Unir todo el texto
    full_text = "".join(full_text_parts)
    
    return ExcelParseResult(
        texto=full_text,
        num_paginas=len(workbook.sheetnames),
        page_offsets=page_offsets,
        sheets_content=sheets_content,
    )


def parse_excel_stream(file_stream: io.BytesIO, filename: str) -> ExcelParseResult:
    """
    Parsea un archivo Excel desde un stream de bytes.
    
    Útil para archivos subidos vía API sin guardar temporalmente en disco.
    
    Args:
        file_stream: Stream de bytes con el contenido del Excel
        filename: Nombre original del archivo (solo para logging)
        
    Returns:
        ExcelParseResult con el contenido extraído
    """
    # openpyxl puede leer directamente desde BytesIO
    workbook = load_workbook(filename=file_stream, data_only=True, read_only=True)
    
    sheets_content: Dict[str, List[List[str]]] = {}
    full_text_parts: List[str] = []
    page_offsets: Dict[int, tuple[int, int]] = {}
    
    current_offset = 0
    
    for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
        sheet: Worksheet = workbook[sheet_name]
        
        # Inicio de offset para esta hoja
        start_offset = current_offset
        
        # Encabezado de hoja
        sheet_header = f"\n{'=' * 80}\n"
        sheet_header += f"HOJA: {sheet_name}\n"
        sheet_header += f"{'=' * 80}\n\n"
        full_text_parts.append(sheet_header)
        current_offset += len(sheet_header)
        
        # Extraer contenido de la hoja
        sheet_rows: List[List[str]] = []
        
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            # Convertir valores de celda a string
            row_values = [
                str(cell) if cell is not None else ""
                for cell in row
            ]
            
            # Filtrar filas completamente vacías
            if any(val.strip() for val in row_values):
                sheet_rows.append(row_values)
                
                # Añadir fila al texto (formato tabular)
                row_text = " | ".join(row_values)
                row_line = f"Fila {row_idx}: {row_text}\n"
                full_text_parts.append(row_line)
                current_offset += len(row_line)
        
        # Guardar contenido estructurado de la hoja
        sheets_content[sheet_name] = sheet_rows
        
        # Fin de offset para esta hoja
        end_offset = current_offset
        page_offsets[sheet_idx] = (start_offset, end_offset)
        
        # Separador entre hojas
        separator = "\n\n"
        full_text_parts.append(separator)
        current_offset += len(separator)
    
    workbook.close()
    
    # Unir todo el texto
    full_text = "".join(full_text_parts)
    
    return ExcelParseResult(
        texto=full_text,
        num_paginas=len(workbook.sheetnames),
        page_offsets=page_offsets,
        sheets_content=sheets_content,
    )


def detect_excel_type(filename: str) -> Optional[str]:
    """
    Detecta si un archivo es un Excel válido por su extensión.
    
    Args:
        filename: Nombre del archivo
        
    Returns:
        "xlsx" o "xls" si es Excel, None si no lo es
    """
    ext = Path(filename).suffix.lower()
    
    if ext == ".xlsx":
        return "xlsx"
    elif ext == ".xls":
        return "xls"
    else:
        return None
