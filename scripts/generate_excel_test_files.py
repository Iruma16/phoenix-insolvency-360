"""
Generador de archivos Excel de prueba para RETAIL DEMO SL.

Genera:
- Balance de Situación en formato Excel
- Cuenta de Pérdidas y Ganancias en formato Excel

Con datos coherentes para probar el parser de Excel.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

OUTPUT_DIR = Path("data/casos_prueba/RETAIL_DEMO_SL")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)


def create_balance_excel():
    """Genera Balance de Situación en formato Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Balance"
    
    # Estilos
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14, color="1a237e")
    subtotal_fill = PatternFill(start_color="e8eaf6", end_color="e8eaf6", fill_type="solid")
    subtotal_font = Font(bold=True)
    total_fill = PatternFill(start_color="c5cae9", end_color="c5cae9", fill_type="solid")
    total_font = Font(bold=True, size=12)
    
    # Título
    ws["A1"] = "RETAIL DEMO SL"
    ws["A1"].font = title_font
    ws["A2"] = "BALANCE DE SITUACIÓN"
    ws["A2"].font = title_font
    ws["A3"] = "A 31 de diciembre de 2023"
    ws["A3"].font = Font(italic=True)
    
    # ACTIVO
    ws["A5"] = "ACTIVO"
    ws["A5"].font = Font(bold=True, size=12)
    
    ws["A7"] = "CONCEPTO"
    ws["B7"] = "IMPORTE (€)"
    ws["A7"].fill = header_fill
    ws["B7"].fill = header_fill
    ws["A7"].font = header_font
    ws["B7"].font = header_font
    
    # Datos de activo
    activo_data = [
        ("A) ACTIVO NO CORRIENTE", 180000, subtotal_fill, subtotal_font),
        ("  I. Inmovilizado intangible", 5000, None, None),
        ("  II. Inmovilizado material", 150000, None, None),
        ("      1. Terrenos y construcciones", 120000, None, None),
        ("      2. Instalaciones técnicas", 30000, None, None),
        ("  III. Inversiones financieras a l/p", 25000, None, None),
        ("", "", None, None),
        ("B) ACTIVO CORRIENTE", 70000, subtotal_fill, subtotal_font),
        ("  I. Existencias", 45000, None, None),
        ("  II. Deudores comerciales", 21500, None, None),
        ("  III. Efectivo y equivalentes", 3500, None, None),
        ("", "", None, None),
        ("TOTAL ACTIVO", 250000, total_fill, total_font),
    ]
    
    row = 8
    for concept, amount, fill, font in activo_data:
        ws[f"A{row}"] = concept
        ws[f"B{row}"] = amount if amount != "" else ""
        if fill:
            ws[f"A{row}"].fill = fill
            ws[f"B{row}"].fill = fill
        if font:
            ws[f"A{row}"].font = font
            ws[f"B{row}"].font = font
        if amount != "":
            ws[f"B{row}"].number_format = '#,##0'
        row += 1
    
    # PASIVO
    ws[f"A{row+1}"] = "PATRIMONIO NETO Y PASIVO"
    ws[f"A{row+1}"].font = Font(bold=True, size=12)
    
    row += 3
    ws[f"A{row}"] = "CONCEPTO"
    ws[f"B{row}"] = "IMPORTE (€)"
    ws[f"A{row}"].fill = header_fill
    ws[f"B{row}"].fill = header_fill
    ws[f"A{row}"].font = header_font
    ws[f"B{row}"].font = header_font
    
    # Datos de pasivo
    pasivo_data = [
        ("A) PATRIMONIO NETO", -230000, subtotal_fill, subtotal_font),
        ("  I. Capital", 60000, None, None),
        ("  II. Reservas", 15000, None, None),
        ("  III. Resultados ejercicio", -60000, None, None),
        ("  IV. Resultados ejercicios anteriores", -245000, None, None),
        ("", "", None, None),
        ("B) PASIVO NO CORRIENTE", 180000, subtotal_fill, subtotal_font),
        ("  I. Deudas a largo plazo", 180000, None, None),
        ("      1. Préstamos entidades crédito", 180000, None, None),
        ("", "", None, None),
        ("C) PASIVO CORRIENTE", 300000, subtotal_fill, subtotal_font),
        ("  I. Deudas a corto plazo", 85000, None, None),
        ("  II. Acreedores comerciales", 105000, None, None),
        ("  III. Deudas con Administraciones", 110000, None, None),
        ("      1. Hacienda Pública", 68000, None, None),
        ("      2. Seguridad Social", 42000, None, None),
        ("", "", None, None),
        ("TOTAL PATRIMONIO NETO Y PASIVO", 250000, total_fill, total_font),
    ]
    
    row += 1
    for concept, amount, fill, font in pasivo_data:
        ws[f"A{row}"] = concept
        ws[f"B{row}"] = amount if amount != "" else ""
        if fill:
            ws[f"A{row}"].fill = fill
            ws[f"B{row}"].fill = fill
        if font:
            ws[f"A{row}"].font = font
            ws[f"B{row}"].font = font
        if amount != "":
            ws[f"B{row}"].number_format = '#,##0'
        row += 1
    
    # Ajustar anchos de columna
    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20
    
    # Guardar
    filename = OUTPUT_DIR / "10_Balance_Situacion_2023.xlsx"
    wb.save(filename)
    print(f"✅ Creado: {filename}")


def create_pyg_excel():
    """Genera Cuenta de PyG en formato Excel."""
    wb = Workbook()
    ws = wb.active
    ws.title = "PyG"
    
    # Estilos
    header_fill = PatternFill(start_color="1a237e", end_color="1a237e", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    title_font = Font(bold=True, size=14, color="1a237e")
    subtotal_fill = PatternFill(start_color="e8eaf6", end_color="e8eaf6", fill_type="solid")
    subtotal_font = Font(bold=True)
    total_fill = PatternFill(start_color="ffebee", end_color="ffebee", fill_type="solid")
    total_font = Font(bold=True, size=12, color="c62828")
    
    # Título
    ws["A1"] = "RETAIL DEMO SL"
    ws["A1"].font = title_font
    ws["A2"] = "CUENTA DE PÉRDIDAS Y GANANCIAS"
    ws["A2"].font = title_font
    ws["A3"] = "Ejercicio 2023"
    ws["A3"].font = Font(italic=True)
    
    ws["A5"] = "CONCEPTO"
    ws["B5"] = "IMPORTE (€)"
    ws["A5"].fill = header_fill
    ws["B5"].fill = header_fill
    ws["A5"].font = header_font
    ws["B5"].font = header_font
    
    # Datos de PyG
    pyg_data = [
        ("", "", None, None),
        ("1. Importe neto de la cifra de negocios", 120000, None, None),
        ("2. Variación de existencias", -15000, None, None),
        ("3. Aprovisionamientos", -65000, None, None),
        ("", "", None, None),
        ("VALOR AÑADIDO (1+2+3)", 40000, subtotal_fill, subtotal_font),
        ("", "", None, None),
        ("4. Gastos de personal", -48000, None, None),
        ("5. Otros gastos de explotación", -32000, None, None),
        ("6. Amortizaciones", -18000, None, None),
        ("7. Deterioros", -5000, None, None),
        ("", "", None, None),
        ("RESULTADO DE EXPLOTACIÓN", -63000, subtotal_fill, subtotal_font),
        ("", "", None, None),
        ("8. Ingresos financieros", 500, None, None),
        ("9. Gastos financieros", -12500, None, None),
        ("", "", None, None),
        ("RESULTADO FINANCIERO", -12000, subtotal_fill, subtotal_font),
        ("", "", None, None),
        ("RESULTADO ANTES DE IMPUESTOS", -75000, subtotal_fill, subtotal_font),
        ("", "", None, None),
        ("10. Impuesto sobre beneficios", 15000, None, None),
        ("", "", None, None),
        ("RESULTADO DEL EJERCICIO", -60000, total_fill, total_font),
    ]
    
    row = 6
    for concept, amount, fill, font in pyg_data:
        ws[f"A{row}"] = concept
        ws[f"B{row}"] = amount if amount != "" else ""
        if fill:
            ws[f"A{row}"].fill = fill
            ws[f"B{row}"].fill = fill
        if font:
            ws[f"A{row}"].font = font
            ws[f"B{row}"].font = font
        if amount != "":
            ws[f"B{row}"].number_format = '#,##0'
        row += 1
    
    # Ajustar anchos de columna
    ws.column_dimensions["A"].width = 50
    ws.column_dimensions["B"].width = 20
    
    # Guardar
    filename = OUTPUT_DIR / "11_Cuenta_PyG_2023.xlsx"
    wb.save(filename)
    print(f"✅ Creado: {filename}")


def main():
    """Genera todos los archivos Excel de prueba."""
    print("\n🚀 Generando archivos Excel para RETAIL DEMO SL\n")
    print("=" * 60)
    
    create_balance_excel()
    create_pyg_excel()
    
    print("=" * 60)
    print(f"\n✅ ARCHIVOS EXCEL GENERADOS\n")
    print(f"📁 Ubicación: {OUTPUT_DIR}")
    print(f"📄 Archivos: 2 documentos Excel")
    print(f"\n🎯 Listos para probar el parser de Excel\n")


if __name__ == "__main__":
    main()
